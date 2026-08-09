import importlib.util
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\hp\Documents\Trainer_Capacity_Hub")
SCRIPT_PATH = ROOT / "backend" / "livrable1.py"
OUT_PATH = ROOT / "backend" / "livrable1_export.xlsx"


spec = importlib.util.spec_from_file_location("livrable1_module", SCRIPT_PATH)
module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(module)


wb = Workbook()
ws = wb.active
ws.title = "Inventaire"

headers = [
    "Catégorie",
    "Élément",
    "Modèle / Enum",
    "Champ(s)",
    "Obligatoire",
    "Facultatif",
    "Sensible",
    "Remarque",
]
ws.append(headers)

rows = [
    ["Référentiel", "RoleCode", "Enum", "FORMATEUR, MANAGER, ADMINISTRATEUR, SUPERADMINISTRATEUR", "Oui", "Non", "Non", "Nomenclature commune des rôles"],
    ["Référentiel", "ActivityTypeCode", "Enum", "ANIMATION, PREPARATION, CORRECTION, REUNION, VISITE, ACCOMPAGNEMENT, GESTION", "Oui", "Non", "Non", "Nomenclature commune des activités"],
    ["Référentiel", "ActivityStatus", "Enum", "VALIDEE, REJETEE, EN_ATTENTE", "Oui", "Non", "Non", "Statut de la charge"],
    ["Référentiel", "LeaveType", "Enum", "CONGE, ABSENCE", "Oui", "Non", "Non", "Disponibilités"],
    ["Référentiel", "UnavailabilityType", "Enum", "MISSION, INDISPONIBILITE_PONCTUELLE, PERIODE_FAIBLE_ACTIVITE, JOUR_FERIE", "Oui", "Non", "Non", "Disponibilités"],
    ["Référentiel", "DeclarationStatus", "Enum", "BROUILLON, SOUMISE, VALIDEE, REJETEE", "Oui", "Non", "Non", "Table propositionnelle"],
    ["Référentiel", "ValidationAction", "Enum", "CREATION, MODIFICATION, VALIDATION, REJET", "Oui", "Non", "Non", "Table propositionnelle"],
    ["Utilisateurs", "User", "BaseModel", "matricule, email, role, active, manager_matricule", "matricule/email/role", "active/manager_matricule", "matricule/email", "Compte d’accès"],
    ["Utilisateurs", "Profile", "BaseModel", "user_matricule, full_name, function, specialty_id, center_id, entry_date", "user_matricule/full_name", "function/specialty_id/center_id/entry_date", "full_name", "Infos métier"],
    ["Charge", "Activity", "BaseModel", "id, trainer_matricule, activity_date, activity_type, duration, program_id, client_id, center_id, status, comment, attachment_ref", "trainer_matricule/activity_date/activity_type/duration", "program_id/client_id/center_id/status/comment/attachment_ref", "trainer_matricule", "Activité de charge"],
    ["Disponibilités", "Leave", "BaseModel", "id, user_matricule, leave_type, start_date, end_date", "user_matricule/leave_type/start_date/end_date", "id", "user_matricule", "Congés et absences"],
    ["Disponibilités", "UnavailabilityPeriod", "BaseModel", "id, user_matricule, unavailability_type, coefficient, start_date, end_date", "unavailability_type/start_date/end_date", "user_matricule/coefficient", "user_matricule", "Missions / indisponibilités / jours fériés"],
    ["Disponibilités", "CalendarExclusion", "BaseModel", "id, label, working_days_excluded, capacity_coefficient, comment, start_date, end_date", "label/working_days_excluded/start_date/end_date", "capacity_coefficient/comment", "Non", "Fenêtres neutralisées"],
    ["Disponibilités", "CapacityTarget", "BaseModel", "id, user_matricule, period_label, theoretical_capacity_days, net_available_capacity_days", "period_label/theoretical_capacity_days", "user_matricule/net_available_capacity_days", "Non", "Capacité cible"],
    ["Capacité", "CapacityRules", "BaseModel", "calendar_days, weekend_days, neutralized_working_days, target_rate, net_capacity_days, computed fields", "calendar_days/weekend_days/neutralized_working_days/target_rate/net_capacity_days", "Non", "Non", "Moteur de calcul"],
]
for row in rows:
    ws.append(row)

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [18, 28, 16, 78, 18, 22, 16, 30]
for idx, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = width

comp = wb.create_sheet("Compatibilite_S1")
comp_headers = ["Demande Semaine 1", "Présent dans livrable1.py", "Statut", "Preuve / correspondance"]
comp.append(comp_headers)
compat_rows = [
    ["Analyser les besoins des formateurs et des managers", "Partiellement", "Partiel", "Le script structure les données métier, mais ne contient pas une analyse fonctionnelle formelle."],
    ["Identifier les principales catégories de données", "Oui", "Compatible", "Référentiels / Utilisateurs / Charge / Disponibilités / Capacité."],
    ["Recenser les informations disponibles dans les fichiers Excel", "Non", "Non compatible", "Aucun import de fichiers Excel source n’est présent dans le script."],
    ["Distinguer données obligatoires et facultatives", "Oui", "Compatible", "Champs required vs default None / default=True."],
    ["Identifier les données sensibles ou confidentielles", "Oui", "Compatible", "Matricule, email, full_name sont signalés comme sensibles."],
    ["Proposer une nomenclature commune", "Oui", "Compatible", "Enums RoleCode, ActivityTypeCode, ActivityStatus, etc."],
    ["Matricule", "Oui", "Compatible", "User.matricule"],
    ["Nom et prénom", "Oui", "Compatible", "Profile.full_name"],
    ["Adresse e-mail", "Oui", "Compatible", "User.email"],
    ["Rôle", "Oui", "Compatible", "User.role / RoleCode"],
    ["Fonction", "Oui", "Compatible", "Profile.function"],
    ["Spécialité", "Oui", "Compatible", "Profile.specialty_id / Specialty"],
    ["Centre de rattachement", "Oui", "Compatible", "Profile.center_id / Center"],
    ["Statut actif ou inactif", "Oui", "Compatible", "User.active"],
    ["Date d’entrée", "Oui", "Compatible", "Profile.entry_date"],
    ["Manager responsable", "Oui", "Compatible", "User.manager_matricule"],
    ["Formateur concerné", "Oui", "Compatible", "Activity.trainer_matricule"],
    ["Date de l’activité", "Oui", "Compatible", "Activity.activity_date"],
    ["Type d’activité", "Oui", "Compatible", "Activity.activity_type / ActivityTypeCode"],
    ["Durée en jours ou en heures", "Oui", "Compatible", "Activity.duration"],
    ["Programme ou formation", "Oui", "Compatible", "Activity.program_id / Program"],
    ["Client", "Oui", "Compatible", "Activity.client_id / Client"],
    ["Centre", "Oui", "Compatible", "Activity.center_id / Center"],
    ["Statut de l’activité", "Oui", "Compatible", "Activity.status / ActivityStatus"],
    ["Commentaire", "Oui", "Compatible", "Activity.comment"],
    ["Justificatif éventuel", "Oui", "Compatible", "Activity.attachment_ref"],
    ["Congés", "Oui", "Compatible", "Leave / LeaveType.CONGE"],
    ["Absences", "Oui", "Compatible", "Leave / LeaveType.ABSENCE"],
    ["Missions", "Oui", "Compatible", "UnavailabilityType.MISSION"],
    ["Indisponibilités", "Oui", "Compatible", "UnavailabilityType.INDISPONIBILITE_PONCTUELLE"],
    ["Périodes de faible activité", "Oui", "Compatible", "UnavailabilityType.PERIODE_FAIBLE_ACTIVITE"],
    ["Jours fériés", "Oui", "Compatible", "UnavailabilityType.JOUR_FERIE"],
    ["Périodes exclues du calcul", "Oui", "Compatible", "CalendarExclusion"],
    ["Capacité théorique", "Oui", "Compatible", "CapacityTarget.theoretical_capacity_days / CapacityRules.net_capacity_days"],
    ["Capacité réellement disponible", "Oui", "Compatible", "CapacityTarget.net_available_capacity_days"],
]
for row in compat_rows:
    comp.append(row)

for cell in comp[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F75B5")
    cell.alignment = Alignment(horizontal="center", vertical="center")

comp_widths = [42, 22, 16, 72]
for idx, width in enumerate(comp_widths, start=1):
    comp.column_dimensions[get_column_letter(idx)].width = width

cap = wb.create_sheet("Regles_Capacite")
cap.append(["Champ", "Valeur", "Explication"])
cap_rows = [
    ["calendar_days", 365, "Jours calendaires 2026"],
    ["weekend_days", 104, "Week-ends"],
    ["neutralized_working_days", 83, "Jours ouvrés neutralisés"],
    ["target_rate", 0.6, "Taux cible animation"],
    ["net_capacity_days", 189.0, "Capacité globale nette"],
    ["working_days", 261, "365 - 104"],
    ["favorable_days", 178, "261 - 83"],
    ["target_days_raw", 106.8, "178 x 0.6"],
    ["target_days_rounded", 107, "Arrondi de pilotage"],
]
for row in cap_rows:
    cap.append(row)

for cell in cap[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="548235")
    cell.alignment = Alignment(horizontal="center", vertical="center")

cap.column_dimensions["A"].width = 28
cap.column_dimensions["B"].width = 16
cap.column_dimensions["C"].width = 48

wb.save(OUT_PATH)
print(OUT_PATH)
